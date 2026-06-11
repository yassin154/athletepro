from flask import Flask, render_template, request, redirect, url_for, session, send_file
from functools import wraps
import psycopg2, psycopg2.extras
import hashlib, os, io, json
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'athletepro_secret_2026')
DATABASE_URL = os.environ.get('DATABASE_URL', 'postgresql://postgres.djdochmdplkfcpwyogou:AthlePro2026!@aws-0-eu-west-1.pooler.supabase.com:6543/postgres')

# ── CHAMPIONNATS PROVISOIRES ──────────────────────────────
CHAMPIONNATS = [
    {'nom': 'Championnat Régional',       'date': '2026-02-15', 'lieu': 'À définir'},
    {'nom': 'Championnat National Indoor','date': '2026-03-10', 'lieu': 'À définir'},
    {'nom': 'Championnat National',       'date': '2026-05-20', 'lieu': 'À définir'},
    {'nom': 'Ch. Arabe Senior',           'date': '2026-06-15', 'lieu': 'À définir'},
    {'nom': 'Jeux Méditerranéens',        'date': '2026-07-01', 'lieu': 'À définir'},
    {'nom': 'Jeux Panarabes',             'date': '2026-08-10', 'lieu': 'À définir'},
    {'nom': 'Jeux Francophonie',          'date': '2026-09-01', 'lieu': 'À définir'},
    {'nom': 'Ch. Afrique Senior',         'date': '2026-10-05', 'lieu': 'À définir'},
    {'nom': 'Ch. Monde Junior',           'date': '2026-11-01', 'lieu': 'À définir'},
    {'nom': 'Ch. Senior Monde 2027',      'date': '2027-08-01', 'lieu': 'À définir'},
    {'nom': 'Ch. Monde Indoor 2028',      'date': '2028-03-01', 'lieu': 'À définir'},
    {'nom': 'JO 2028',                    'date': '2028-07-15', 'lieu': 'Los Angeles'},
]

# ── LOAD JSON DATA ─────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

def load_athletes_json():
    path = os.path.join(BASE_DIR, 'athletes.json')
    with open(path, encoding='utf-8') as f:
        return json.load(f)

def load_resultats_json():
    path = os.path.join(BASE_DIR, 'resultats.json')
    with open(path, encoding='utf-8') as f:
        return json.load(f)

# Cache in memory at startup
try:
    with open(os.path.join(BASE_DIR, 'epreuves_ref.json'), encoding='utf-8') as _f:
        EPREUVES_REF = json.load(_f)
    with open(os.path.join(BASE_DIR, 'minimas.json'), encoding='utf-8') as _f:
        MINIMAS = json.load(_f)
    with open(os.path.join(BASE_DIR, 'epreuve_to_specialite.json'), encoding='utf-8') as _f:
        EPREUVE_TO_SPECIALITE = json.load(_f)
    ATHLETES_JSON = load_athletes_json()
    RESULTATS_JSON = load_resultats_json()
    # Build lookup by licence
    ATHLETES_BY_LICENCE = {a['licence']: a for a in ATHLETES_JSON}
    RESULTATS_BY_LICENCE = {}
    for r in RESULTATS_JSON:
        lic = r['licence']
        if lic not in RESULTATS_BY_LICENCE:
            RESULTATS_BY_LICENCE[lic] = []
        RESULTATS_BY_LICENCE[lic].append(r)
    print(f"✅ JSON chargé: {len(ATHLETES_JSON)} athlètes, {len(RESULTATS_JSON)} résultats")
except Exception as e:
    print(f"⚠️ Erreur chargement JSON: {e}")
    ATHLETES_JSON = []
    RESULTATS_JSON = []
    ATHLETES_BY_LICENCE = {}
    RESULTATS_BY_LICENCE = {}
    EPREUVES_REF = {}
    EPREUVE_TO_SPECIALITE = {}
    MINIMAS = {}

# ── DB ─────────────────────────────────────────────────────
def get_db():
    conn = psycopg2.connect(DATABASE_URL, sslmode='require', connect_timeout=10)
    conn.autocommit = False
    return conn

def q(conn, sql, params=(), one=False):
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(sql, params)
    return cur.fetchone() if one else cur.fetchall()

def ex(conn, sql, params=()):
    cur = conn.cursor()
    cur.execute(sql, params)
    conn.commit()

def ex_ret(conn, sql, params=()):
    cur = conn.cursor()
    cur.execute(sql, params)
    conn.commit()
    return cur.fetchone()[0] if cur.description else None

def hash_pw(pw):
    return hashlib.sha256(pw.encode()).hexdigest()

def calc_age(dob_str):
    if not dob_str: return None
    try:
        dob = datetime.strptime(str(dob_str)[:10], '%Y-%m-%d')
        return (datetime.now() - dob).days // 365
    except:
        return None

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get('role') != 'admin':
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated

def init_db():
    conn = get_db()

    ex(conn, '''CREATE TABLE IF NOT EXISTS users (
        id SERIAL PRIMARY KEY,
        username TEXT UNIQUE NOT NULL,
        password TEXT NOT NULL,
        full_name TEXT NOT NULL,
        role TEXT NOT NULL DEFAULT 'coach'
    )''')

    ex(conn, '''CREATE TABLE IF NOT EXISTS athletes (
        id SERIAL PRIMARY KEY,
        centre TEXT, coach_id INTEGER,
        nom_prenom TEXT NOT NULL,
        date_naissance TEXT, age INTEGER,
        numero_licence TEXT UNIQUE,
        categorie TEXT, sexe TEXT,
        specialite TEXT, epreuves TEXT,
        club TEXT, statut TEXT,
        date_integration TEXT,
        classe TEXT DEFAULT 'Autre',
        FOREIGN KEY(coach_id) REFERENCES users(id)
    )''')

    ex(conn, '''CREATE TABLE IF NOT EXISTS objectifs (
        id SERIAL PRIMARY KEY,
        athlete_id INTEGER NOT NULL,
        epreuve TEXT NOT NULL,
        saison TEXT DEFAULT '2025/2026',
        objectif_chrono TEXT,
        statut TEXT DEFAULT 'en_attente',
        soumis_le TEXT,
        valide_le TEXT,
        FOREIGN KEY(athlete_id) REFERENCES athletes(id)
    )''')
    try:
        ex(conn, "ALTER TABLE objectifs ADD COLUMN IF NOT EXISTS saison TEXT DEFAULT '2025/2026'")
    except:
        pass

    ex(conn, '''CREATE TABLE IF NOT EXISTS objectifs_champ (
        id SERIAL PRIMARY KEY,
        athlete_id INTEGER NOT NULL,
        epreuve TEXT NOT NULL,
        championnat TEXT NOT NULL,
        participe INTEGER DEFAULT 0,
        objectif TEXT,
        statut TEXT DEFAULT 'en_attente',
        soumis_le TEXT,
        valide_le TEXT,
        FOREIGN KEY(athlete_id) REFERENCES athletes(id)
    )''')

    ex(conn, '''CREATE TABLE IF NOT EXISTS resultats (
        id SERIAL PRIMARY KEY,
        athlete_id INTEGER NOT NULL,
        epreuve TEXT NOT NULL,
        nom_competition TEXT,
        lieu TEXT,
        date_competition TEXT,
        performance TEXT,
        classement TEXT,
        statut TEXT DEFAULT 'en_attente',
        soumis_le TEXT,
        valide_le TEXT,
        FOREIGN KEY(athlete_id) REFERENCES athletes(id)
    )''')
    try:
        ex(conn, "ALTER TABLE resultats ADD COLUMN IF NOT EXISTS classement TEXT")
    except:
        pass

    ex(conn, '''CREATE TABLE IF NOT EXISTS objectifs_perf_res (
        id SERIAL PRIMARY KEY,
        athlete_id INTEGER NOT NULL,
        source TEXT NOT NULL,
        source_id INTEGER NOT NULL,
        epreuve TEXT,
        objectif_perf TEXT,
        statut TEXT DEFAULT 'en_attente',
        soumis_le TEXT,
        valide_le TEXT,
        UNIQUE(athlete_id, source, source_id),
        FOREIGN KEY(athlete_id) REFERENCES athletes(id)
    )''')

    ex(conn, '''CREATE TABLE IF NOT EXISTS historique (
        id SERIAL PRIMARY KEY,
        athlete_id INTEGER NOT NULL,
        epreuve TEXT NOT NULL,
        saison TEXT,
        competition TEXT,
        lieu TEXT,
        date_competition TEXT,
        performance TEXT,
        FOREIGN KEY(athlete_id) REFERENCES athletes(id)
    )''')

    # Seed admin
    try:
        ex(conn, "ALTER TABLE athletes ADD COLUMN IF NOT EXISTS classe TEXT DEFAULT 'Autre'")
    except: pass

    ex(conn, '''INSERT INTO users (username,password,full_name,role)
                VALUES (%s,%s,%s,%s) ON CONFLICT (username) DO NOTHING''',
       ('admin', hash_pw('Admin2026!'), 'Yassine Bouta', 'admin'))

    coaches = [
        ('AIT EL HAJ KARIM',  'ait.el.haj.karim'),
        ('ALI EZZINE',         'ali.ezzine'),
        ('BAAKIL SOUFIANE',    'baakil.soufiane'),
        ('BELMRHAR HICHAM',    'belmrhar.hicham'),
        ('BOUKRAA ABDELLAH',   'boukraa.abdellah'),
        ('CHERQAOUI',          'cherqaoui'),
        ('ECHAFIYI TOUFIK',    'echafiyi.toufik'),
        ('ELMOUADEN',          'elmouaden'),
        ('ITFAN LAHCEN',       'itfan.lahcen'),
        ('KABBOU BRAHIM',      'kabbou.brahim'),
        ('KAHLAOUI MAROUANE',  'kahlaoui.marouane'),
        ('KARIM TELMCANI',     'karim.telmcani'),
        ('MAHJOUR AHMED',      'mahjour.ahmed'),
        ('MOUHCINE JAMAL',     'mouhcine.jamal'),
        ('NABAOUI HAFID',      'nabaoui.hafid'),
        ('OUKHBACH HASSAN',    'oukhbach.hassan'),
        ('OUZLIM MOHAMED',     'ouzlim.mohamed'),
        ('ROUAS HATIM',        'rouas.hatim'),
        ('SAKAH FADOUA',       'sakah.fadoua'),
        ('SEKOURI JALAL',      'sekouri.jalal'),
        ('SKAH KHALID',        'skah.khalid'),
        ('ZERAIDI EL MEHDI',   'zeraidi.elmehdi'),
    ]
    try:
        ex(conn, "ALTER TABLE athletes ADD COLUMN IF NOT EXISTS classe TEXT DEFAULT 'Autre'")
    except: pass

    for full_name, username in coaches:
        ex(conn, '''INSERT INTO users (username,password,full_name,role)
                    VALUES (%s,%s,%s,%s) ON CONFLICT (username) DO NOTHING''',
           (username, hash_pw('Coach2026'), full_name, 'coach'))

    # Seed athletes from JSON
    for a in ATHLETES_JSON:
        coach = q(conn, "SELECT id FROM users WHERE full_name=%s", (a['entraineur'],), one=True)
        if coach:
            epreuves_str = '/'.join(a.get('epreuves', []))
            ex(conn, '''INSERT INTO athletes
                (centre,coach_id,nom_prenom,date_naissance,age,numero_licence,categorie,sexe,specialite,epreuves,club,statut,date_integration,classe)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT (numero_licence) DO UPDATE SET
                    club=EXCLUDED.club, epreuves=EXCLUDED.epreuves,
                    age=EXCLUDED.age, specialite=EXCLUDED.specialite,
                    classe=EXCLUDED.classe''',
                (a['crf'], coach['id'], a['nom'], a['date_naissance'], a['age'],
                 a['licence'], a['categorie'], a['sexe'], a['specialite'],
                 epreuves_str, a['club'], a['statut'], a['integration'],
                 a.get('classe', 'Autre')))

    conn.close()
    print("✅ DB initialisée")

# ── ROUTES ────────────────────────────────────────────────

@app.route('/')
def index():
    if 'user_id' in session:
        if session.get('role') == 'admin':
            return redirect(url_for('admin_dashboard'))
        return redirect(url_for('guide'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        conn = get_db()
        user = q(conn, "SELECT * FROM users WHERE username=%s AND password=%s",
                 (username, hash_pw(password)), one=True)
        conn.close()
        if user:
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['full_name'] = user['full_name']
            session['role'] = user['role']
            if user['role'] == 'admin':
                return redirect(url_for('admin_dashboard'))
            return redirect(url_for('guide'))
        error = "Identifiant ou mot de passe incorrect"
    return render_template('login.html', error=error)

@app.route('/guide')
@login_required
def guide():
    if session.get('role') == 'admin':
        return redirect(url_for('admin_dashboard'))
    return render_template('guide.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    conn = get_db()
    if session['role'] == 'admin':
        athletes = q(conn, """
            SELECT a.*, u.full_name as coach_name
            FROM athletes a JOIN users u ON a.coach_id=u.id
            ORDER BY
                CASE a.categorie
                    WHEN 'M1' THEN 1 WHEN 'M2' THEN 2
                    WHEN 'C1' THEN 3 WHEN 'C2' THEN 4
                    WHEN 'J1' THEN 5 WHEN 'J2' THEN 6
                    WHEN 'S'  THEN 7 ELSE 8 END,
                a.nom_prenom
        """)
    else:
        athletes = q(conn, """
            SELECT a.*, u.full_name as coach_name
            FROM athletes a JOIN users u ON a.coach_id=u.id
            WHERE a.coach_id=%s
            ORDER BY
                CASE a.categorie
                    WHEN 'M1' THEN 1 WHEN 'M2' THEN 2
                    WHEN 'C1' THEN 3 WHEN 'C2' THEN 4
                    WHEN 'J1' THEN 5 WHEN 'J2' THEN 6
                    WHEN 'S'  THEN 7 ELSE 8 END,
                a.nom_prenom
        """, (session['user_id'],))
    conn.close()

    # Group by category
    cats = {'M1':[],'M2':[],'C1':[],'C2':[],'J1':[],'J2':[],'S':[],'Autre':[]}
    for a in athletes:
        cat = a['categorie'] if a['categorie'] in cats else 'Autre'
        cats[cat].append(a)

    return render_template('dashboard.html', categories=cats, athletes=athletes)

@app.route('/athlete/<int:aid>')
@login_required
def athlete(aid):
    conn = get_db()
    a = q(conn, """SELECT a.*, u.full_name as coach_name
                   FROM athletes a JOIN users u ON a.coach_id=u.id
                   WHERE a.id=%s""", (aid,), one=True)
    if not a:
        conn.close()
        return "Athlète introuvable", 404

    # Authorization check
    if session['role'] != 'admin' and a['coach_id'] != session['user_id']:
        conn.close()
        return "Accès refusé", 403

    licence = a['numero_licence']

    # Get reference epreuves from JSON (based on cat+sex)
    a_json = ATHLETES_BY_LICENCE.get(licence, {})
    epreuves_ref_list = a_json.get('epreuves_ref', [])
    specialite_norm = a_json.get('specialite_norm', a.get('specialite',''))

    # Epreuve -> specialite mapping
    ep_to_spec = EPREUVE_TO_SPECIALITE

    # Get epreuves from DB + JSON (merged)
    db_epreuves = set(e.strip() for e in (a['epreuves'] or '').split('/') if e.strip())
    json_epreuves = set(ATHLETES_BY_LICENCE.get(licence, {}).get('epreuves', []))
    all_epreuves = sorted(db_epreuves | json_epreuves)

    # Historical results from JSON
    hist = RESULTATS_BY_LICENCE.get(licence, [])
    # Group by saison dynamically
    from collections import defaultdict as _dd
    hist_by_saison = _dd(list)
    for r in hist:
        hist_by_saison[r.get('saison','')].append(r)
    # Keep named ones for template backward compat
    hist_2425 = hist_by_saison.get('2024/2025', [])
    hist_2526 = hist_by_saison.get('2025/2026', [])
    # All other saisons (historical)
    hist_autres = {s: v for s, v in sorted(hist_by_saison.items())
                   if s not in ('2024/2025','2025/2026') and s}

    # Objectifs from DB
    objectifs = q(conn, "SELECT * FROM objectifs WHERE athlete_id=%s ORDER BY saison,epreuve", (aid,))
    obj_map_2425 = {o['epreuve']: o for o in objectifs if o.get('saison') == '2024/2025' and o['statut'] == 'validé'}
    obj_map_2526 = {o['epreuve']: o for o in objectifs if o.get('saison') == '2025/2026' and o['statut'] == 'validé'}

    # Current season submitted results
    resultats = q(conn, "SELECT * FROM resultats WHERE athlete_id=%s AND statut='validé' ORDER BY date_competition DESC", (aid,))

    # Championships
    champs = q(conn, "SELECT * FROM objectifs_champ WHERE athlete_id=%s", (aid,))
    champ_map = {}
    for c in champs:
        key = f"{c['epreuve']}|{c['championnat']}"
        champ_map[key] = c
    validated_champs = q(conn, "SELECT * FROM objectifs_champ WHERE athlete_id=%s AND statut='validé'", (aid,))

    # Perf objectives on historical results
    perf_objs = q(conn, "SELECT * FROM objectifs_perf_res WHERE athlete_id=%s AND statut='validé'", (aid,))
    perf_obj_map = {}
    for p in perf_objs:
        perf_obj_map[(p['source'], p['source_id'])] = p

    conn.close()

    # Build chart data for evolution tab — include ALL historical results
    all_hist = []
    all_sources = list(hist) + list(resultats)  # hist = all JSON results, resultats = DB submitted
    for r in all_sources:
        perf_str = r.get('performance') or r.get('resultat') or ''
        if not perf_str or perf_str in ('nan', 'None', '—', ''):
            continue
        all_hist.append({
            'date': r.get('date') or r.get('date_competition') or '',
            'epreuve': r.get('epreuve', ''),
            'competition': r.get('competition') or r.get('nom_competition') or '',
            'lieu': r.get('lieu', ''),
            'performance': perf_str,
            'classement': str(r.get('classement', '')) if r.get('classement') else '',
            'saison': r.get('saison') or r.get('saison_label') or '2025/2026',
            'source': 'db' if 'athlete_id' in r else 'json'
        })

    all_hist.sort(key=lambda x: x['date'] or '')
    # Only keep real epreuves (those in our mapping) for chart, exclude raw specialite names
    specialite_names = set(EPREUVE_TO_SPECIALITE.values())
    all_epreuves_chart = sorted(set(
        r['epreuve'] for r in all_hist
        if r['epreuve'] and r['epreuve'] not in specialite_names
        and r['epreuve'] in EPREUVE_TO_SPECIALITE
    ))
    # Also add any epreuve not in mapping but looks like a real one (has digits)
    import re as _re
    for r in all_hist:
        ep = r.get('epreuve','')
        if ep and ep not in specialite_names and ep not in all_epreuves_chart:
            if _re.search(r'[0-9]', ep):
                all_epreuves_chart.append(ep)
    all_epreuves_chart = sorted(set(all_epreuves_chart))

    # Build res_by_epreuve for template
    res_by_epreuve = {}
    for r in resultats:
        ep = r['epreuve']
        if ep not in res_by_epreuve:
            res_by_epreuve[ep] = []
        res_by_epreuve[ep].append(r)

    def obj_map_to_json(m):
        return {k: dict(v) for k, v in m.items()}

    # Build all-saisons obj map for JS chart: {saison: {epreuve: chrono}}
    all_obj_map_json = {}
    for o in objectifs:
        if o.get('statut') == 'validé' and o.get('epreuve') and o.get('objectif_chrono'):
            s = o.get('saison') or '2025/2026'
            if s not in all_obj_map_json:
                all_obj_map_json[s] = {}
            all_obj_map_json[s][o['epreuve']] = o['objectif_chrono']

    # Get active tab from URL param (e.g. after form submit)
    active_tab = request.args.get('tab', 'tab-info')

    return render_template('athlete.html',
        a=a, ath=a,
        all_epreuves=all_epreuves,
        historique_2425=hist_2425,
        historique_2526=hist_2526,
        hist_autres=hist_autres,
        res_by_epreuve=res_by_epreuve,
        obj_map_2425=obj_map_2425,
        obj_map_2526=obj_map_2526,
        obj_map_2425_json=obj_map_to_json(obj_map_2425),
        obj_map_2526_json=obj_map_to_json(obj_map_2526),
        all_obj_map_json=all_obj_map_json,
        objectifs=objectifs,
        obj_champ_map=champ_map,
        validated_champs=validated_champs,
        perf_obj_map=perf_obj_map,
        championnats=CHAMPIONNATS,
        chart_data=all_hist,
        chart_epreuves=all_epreuves_chart,
        active_tab=active_tab,
        epreuves_ref=epreuves_ref_list,
        specialite_norm=specialite_norm,
        epreuves_ref_all=EPREUVES_REF,
        ep_to_spec=ep_to_spec,
    )

@app.route('/athlete/<int:aid>/add_epreuve', methods=['POST'])
@login_required
def add_epreuve(aid):
    conn = get_db()
    a = q(conn, "SELECT * FROM athletes WHERE id=%s", (aid,), one=True)
    if not a or (session['role'] != 'admin' and a['coach_id'] != session['user_id']):
        conn.close()
        return "Accès refusé", 403

    new_ep = request.form.get('new_epreuve', '').strip()
    if new_ep:
        current = a['epreuves'] or ''
        eps = [e.strip() for e in current.split('/') if e.strip()]
        if new_ep not in eps:
            eps.append(new_ep)
            ex(conn, "UPDATE athletes SET epreuves=%s WHERE id=%s", ('/'.join(eps), aid))

    conn.close()
    return redirect(url_for('athlete', aid=aid))

@app.route('/athlete/<int:aid>/submit_objectifs', methods=['POST'])
@login_required
def submit_objectifs(aid):
    conn = get_db()
    a = q(conn, "SELECT * FROM athletes WHERE id=%s", (aid,), one=True)
    if not a or (session['role'] != 'admin' and a['coach_id'] != session['user_id']):
        conn.close()
        return "Accès refusé", 403

    saison = request.form.get('saison', '2025/2026')
    epreuves = request.form.getlist('epreuve[]')
    objectifs = request.form.getlist('objectif_chrono[]')
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    for ep, obj in zip(epreuves, objectifs):
        if obj.strip():
            existing = q(conn, "SELECT id FROM objectifs WHERE athlete_id=%s AND epreuve=%s AND saison=%s",
                         (aid, ep, saison), one=True)
            if existing:
                ex(conn, "UPDATE objectifs SET objectif_chrono=%s, statut='en_attente', soumis_le=%s WHERE id=%s",
                   (obj, now, existing['id']))
            else:
                ex(conn, """INSERT INTO objectifs (athlete_id,epreuve,saison,objectif_chrono,statut,soumis_le)
                            VALUES (%s,%s,%s,%s,'en_attente',%s)""", (aid, ep, saison, obj, now))

    conn.close()
    return redirect(url_for('athlete', aid=aid) + '?tab=objectifs')

@app.route('/athlete/<int:aid>/submit_resultat', methods=['POST'])
@login_required
def submit_resultat(aid):
    conn = get_db()
    a = q(conn, "SELECT * FROM athletes WHERE id=%s", (aid,), one=True)
    if not a or (session['role'] != 'admin' and a['coach_id'] != session['user_id']):
        conn.close()
        return "Accès refusé", 403

    epreuve = request.form.get('epreuve', '').strip()
    if epreuve == '__new__':
        epreuve = request.form.get('new_epreuve', '').strip()
        if epreuve:
            current = a['epreuves'] or ''
            eps = [e.strip() for e in current.split('/') if e.strip()]
            if epreuve not in eps:
                eps.append(epreuve)
                ex(conn, "UPDATE athletes SET epreuves=%s WHERE id=%s", ('/'.join(eps), aid))

    nom_competition = request.form.get('nom_competition', '').strip()
    lieu = request.form.get('lieu', '').strip()
    date_competition = request.form.get('date_competition', '').strip()
    performance = request.form.get('performance', '').strip()
    classement = request.form.get('classement', '').strip()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    if epreuve and performance:
        rid = ex_ret(conn, """INSERT INTO resultats (athlete_id,epreuve,nom_competition,lieu,date_competition,performance,classement,statut,soumis_le)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,'en_attente',%s) RETURNING id""",
           (aid, epreuve, nom_competition, lieu, date_competition, performance, classement, now))
        # If inline objectif_perf provided, save it too
        obj_perf = request.form.get('objectif_perf_inline', '').strip()
        if obj_perf and rid:
            ex(conn, """INSERT INTO objectifs_perf_res (athlete_id,source,source_id,epreuve,objectif_perf,statut,soumis_le)
                        VALUES (%s,'resultat',%s,%s,%s,'en_attente',%s)
                        ON CONFLICT (athlete_id,source,source_id) DO UPDATE
                        SET objectif_perf=%s, statut='en_attente', soumis_le=%s""",
               (aid, rid, epreuve, obj_perf, now, obj_perf, now))

    conn.close()
    return redirect(url_for('athlete', aid=aid))

@app.route('/athlete/<int:aid>/submit_perf_obj', methods=['POST'])
@login_required
def submit_perf_obj(aid):
    conn = get_db()
    a = q(conn, "SELECT * FROM athletes WHERE id=%s", (aid,), one=True)
    if not a or (session['role'] != 'admin' and a['coach_id'] != session['user_id']):
        conn.close()
        return "Accès refusé", 403

    source = request.form.get('source')
    source_id = request.form.get('source_id')
    epreuve = request.form.get('epreuve', '')
    objectif_perf = request.form.get('objectif_perf', '').strip()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    if objectif_perf:
        ex(conn, """INSERT INTO objectifs_perf_res (athlete_id,source,source_id,epreuve,objectif_perf,statut,soumis_le)
                    VALUES (%s,%s,%s,%s,%s,'en_attente',%s)
                    ON CONFLICT (athlete_id,source,source_id) DO UPDATE
                    SET objectif_perf=%s, statut='en_attente', soumis_le=%s""",
           (aid, source, source_id, epreuve, objectif_perf, now, objectif_perf, now))

    conn.close()
    return redirect(url_for('athlete', aid=aid))

@app.route('/athlete/<int:aid>/submit_champ', methods=['POST'])
@login_required
def submit_champ(aid):
    conn = get_db()
    a = q(conn, "SELECT * FROM athletes WHERE id=%s", (aid,), one=True)
    if not a or (session['role'] != 'admin' and a['coach_id'] != session['user_id']):
        conn.close()
        return "Accès refusé", 403

    epreuve = request.form.get('epreuve', '').strip()
    championnat = request.form.get('championnat', '').strip()
    participe = 1 if request.form.get('participe') else 0
    objectif = request.form.get('objectif', '').strip()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    if epreuve and championnat:
        existing = q(conn, "SELECT id FROM objectifs_champ WHERE athlete_id=%s AND epreuve=%s AND championnat=%s",
                     (aid, epreuve, championnat), one=True)
        if existing:
            ex(conn, """UPDATE objectifs_champ SET participe=%s, objectif=%s, statut='en_attente', soumis_le=%s
                        WHERE id=%s""", (participe, objectif, now, existing['id']))
        else:
            ex(conn, """INSERT INTO objectifs_champ (athlete_id,epreuve,championnat,participe,objectif,statut,soumis_le)
                        VALUES (%s,%s,%s,%s,%s,'en_attente',%s)""",
               (aid, epreuve, championnat, participe, objectif, now))

    conn.close()
    return redirect(url_for('athlete', aid=aid))

@app.route('/athlete/<int:aid>/submit_champ_bulk', methods=['POST'])
@login_required
def submit_champ_bulk(aid):
    conn = get_db()
    a = q(conn, "SELECT * FROM athletes WHERE id=%s", (aid,), one=True)
    if not a or (session['role'] != 'admin' and a['coach_id'] != session['user_id']):
        conn.close()
        return "Accès refusé", 403

    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    epreuves = request.form.getlist('epreuve[]')

    for epr_idx, epreuve in enumerate(epreuves, 1):
        champ_count = int(request.form.get(f'champ_count_{epr_idx}', 0))
        for c_idx in range(1, champ_count + 1):
            champ_nom = request.form.get(f'champ_nom_{epr_idx}_{c_idx}', '').strip()
            participe = 1 if request.form.get(f'champ_participe_{epr_idx}_{c_idx}') else 0
            objectif = request.form.get(f'champ_objectif_{epr_idx}_{c_idx}', '').strip()
            if not champ_nom:
                continue
            existing = q(conn, "SELECT id FROM objectifs_champ WHERE athlete_id=%s AND epreuve=%s AND championnat=%s",
                         (aid, epreuve, champ_nom), one=True)
            if participe or existing:
                upd = "UPDATE objectifs_champ SET participe=%s, objectif=%s, statut='en_attente', soumis_le=%s WHERE id=%s"
                ins = "INSERT INTO objectifs_champ (athlete_id,epreuve,championnat,participe,objectif,statut,soumis_le) VALUES (%s,%s,%s,%s,%s,'en_attente',%s)"
                if existing:
                    ex(conn, upd, (participe, objectif, now, existing['id']))
                else:
                    ex(conn, ins, (aid, epreuve, champ_nom, participe, objectif, now))

    conn.close()
    return redirect(url_for('athlete', aid=aid))

@app.route('/admin')
@login_required
@admin_required
def admin_dashboard():
    conn = get_db()

    # Stats
    total_athletes = q(conn, "SELECT COUNT(*) as n FROM athletes", one=True)['n']
    total_coaches = q(conn, "SELECT COUNT(*) as n FROM users WHERE role='coach'", one=True)['n']

    pending_obj = q(conn, """
        SELECT o.*, a.nom_prenom, a.categorie, u.full_name as coach_name
        FROM objectifs o
        JOIN athletes a ON o.athlete_id=a.id
        JOIN users u ON a.coach_id=u.id
        WHERE o.statut='en_attente'
        ORDER BY o.soumis_le DESC
    """)

    pending_res = q(conn, """
        SELECT r.*, a.nom_prenom, a.categorie, u.full_name as coach_name
        FROM resultats r
        JOIN athletes a ON r.athlete_id=a.id
        JOIN users u ON a.coach_id=u.id
        WHERE r.statut='en_attente'
        ORDER BY r.soumis_le DESC
    """)

    pending_champ = q(conn, """
        SELECT c.*, a.nom_prenom, a.categorie, u.full_name as coach_name
        FROM objectifs_champ c
        JOIN athletes a ON c.athlete_id=a.id
        JOIN users u ON a.coach_id=u.id
        WHERE c.statut='en_attente'
        ORDER BY c.soumis_le DESC
    """)

    pending_perf = q(conn, """
        SELECT p.*, a.nom_prenom, a.categorie, u.full_name as coach_name
        FROM objectifs_perf_res p
        JOIN athletes a ON p.athlete_id=a.id
        JOIN users u ON a.coach_id=u.id
        WHERE p.statut='en_attente'
        ORDER BY p.soumis_le DESC
    """)

    all_athletes = q(conn, """
        SELECT a.*, u.full_name as coach_name
        FROM athletes a JOIN users u ON a.coach_id=u.id
        ORDER BY a.categorie, a.nom_prenom
    """)

    coaches = q(conn, "SELECT * FROM users WHERE role='coach' ORDER BY full_name")

    validated_obj = q(conn, """
        SELECT o.*, a.nom_prenom, a.centre, a.categorie, u.full_name as coach_name
        FROM objectifs o
        JOIN athletes a ON o.athlete_id=a.id
        JOIN users u ON a.coach_id=u.id
        WHERE o.statut='validé'
        ORDER BY o.valide_le DESC
    """)

    validated_res = q(conn, """
        SELECT r.*, a.nom_prenom, a.centre, a.categorie, u.full_name as coach_name
        FROM resultats r
        JOIN athletes a ON r.athlete_id=a.id
        JOIN users u ON a.coach_id=u.id
        WHERE r.statut='validé'
        ORDER BY r.valide_le DESC
    """)

    conn.close()

    stats = {
        'total_athletes': total_athletes,
        'total_coaches': total_coaches,
        'en_attente': len(pending_obj) + len(pending_res) + len(pending_champ) + len(pending_perf),
        'validated': 0  # filled after query
    }

    return render_template('admin.html',
        stats=stats,
        pending_obj=pending_obj,
        pending_res=pending_res,
        pending_champ=pending_champ,
        pending_perf=pending_perf,
        athletes=all_athletes,
        coaches=coaches,
        validated_obj=validated_obj,
        validated_res=validated_res,
    )

@app.route('/admin/valider_obj/<int:oid>', methods=['GET','POST'])
@login_required
@admin_required
def valider_objectif(oid):
    conn = get_db()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ex(conn, "UPDATE objectifs SET statut='validé', valide_le=%s WHERE id=%s", (now, oid))
    conn.close()
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/rejeter_obj/<int:oid>', methods=['GET','POST'])
@login_required
@admin_required
def rejeter_objectif(oid):
    conn = get_db()
    ex(conn, "UPDATE objectifs SET statut='rejeté' WHERE id=%s", (oid,))
    conn.close()
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/valider_res/<int:rid>', methods=['GET','POST'])
@login_required
@admin_required
def valider_resultat(rid):
    conn = get_db()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ex(conn, "UPDATE resultats SET statut='validé', valide_le=%s WHERE id=%s", (now, rid))
    conn.close()
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/rejeter_res/<int:rid>', methods=['GET','POST'])
@login_required
@admin_required
def rejeter_resultat(rid):
    conn = get_db()
    ex(conn, "UPDATE resultats SET statut='rejeté' WHERE id=%s", (rid,))
    conn.close()
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/valider_perf/<int:pid>', methods=['GET','POST'])
@login_required
@admin_required
def valider_perf_obj(pid):
    conn = get_db()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ex(conn, "UPDATE objectifs_perf_res SET statut='validé', valide_le=%s WHERE id=%s", (now, pid))
    conn.close()
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/rejeter_perf/<int:pid>', methods=['GET','POST'])
@login_required
@admin_required
def rejeter_perf_obj(pid):
    conn = get_db()
    ex(conn, "UPDATE objectifs_perf_res SET statut='rejeté' WHERE id=%s", (pid,))
    conn.close()
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/valider_champ/<int:cid>', methods=['GET','POST'])
@login_required
@admin_required
def valider_champ(cid):
    conn = get_db()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ex(conn, "UPDATE objectifs_champ SET statut='validé', valide_le=%s WHERE id=%s", (now, cid))
    conn.close()
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/rejeter_champ/<int:cid>', methods=['GET','POST'])
@login_required
@admin_required
def rejeter_champ(cid):
    conn = get_db()
    ex(conn, "UPDATE objectifs_champ SET statut='rejeté' WHERE id=%s", (cid,))
    conn.close()
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/add_coach', methods=['POST'])
@login_required
@admin_required
def add_coach():
    full_name = request.form.get('full_name', '').strip().upper()
    username  = request.form.get('username', '').strip().lower()
    password  = request.form.get('password', 'Coach2026').strip()
    if full_name and username:
        conn = get_db()
        ex(conn, '''INSERT INTO users (username,password,full_name,role)
                    VALUES (%s,%s,%s,'coach') ON CONFLICT (username) DO NOTHING''',
           (username, hash_pw(password), full_name))
        conn.close()
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/supprimer_coach/<int:uid>', methods=['POST'])
@login_required
@admin_required
def supprimer_coach(uid):
    conn = get_db()
    ex(conn, "DELETE FROM users WHERE id=%s AND role='coach'", (uid,))
    conn.close()
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/supprimer_obj/<int:oid>', methods=['POST'])
@login_required
@admin_required
def supprimer_obj(oid):
    conn = get_db()
    ex(conn, "DELETE FROM objectifs WHERE id=%s", (oid,))
    conn.close()
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/supprimer_res/<int:rid>', methods=['POST'])
@login_required
@admin_required
def supprimer_res(rid):
    conn = get_db()
    ex(conn, "DELETE FROM resultats WHERE id=%s", (rid,))
    conn.close()
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/coach/<int:uid>/dashboard')
@login_required
@admin_required
def admin_coach_dashboard(uid):
    """Admin views a coach's dashboard without login — read-only impersonation."""
    conn = get_db()
    coach = q(conn, "SELECT * FROM users WHERE id=%s AND role='coach'", (uid,), one=True)
    if not coach:
        conn.close()
        return "Entraineur introuvable", 404

    athletes = q(conn, """
        SELECT a.*, u.full_name as coach_name
        FROM athletes a JOIN users u ON a.coach_id=u.id
        WHERE a.coach_id=%s
        ORDER BY
            CASE a.categorie
                WHEN 'M1' THEN 1 WHEN 'M2' THEN 2
                WHEN 'C1' THEN 3 WHEN 'C2' THEN 4
                WHEN 'J1' THEN 5 WHEN 'J2' THEN 6
                WHEN 'S'  THEN 7 ELSE 8 END,
            a.nom_prenom
    """, (uid,))
    conn.close()

    cats = {'M1':[],'M2':[],'C1':[],'C2':[],'J1':[],'J2':[],'S':[],'Autre':[]}
    for a in athletes:
        cat = a['categorie'] if a['categorie'] in cats else 'Autre'
        cats[cat].append(a)

    return render_template('dashboard.html',
        categories=cats,
        athletes=athletes,
        admin_view_coach=coach['full_name'],
    )

@app.route('/admin/athlete/<int:aid>')
@login_required
@admin_required
def admin_athlete(aid):
    conn = get_db()
    a = q(conn, """SELECT a.*, u.full_name as coach_name
                   FROM athletes a JOIN users u ON a.coach_id=u.id
                   WHERE a.id=%s""", (aid,), one=True)
    if not a:
        conn.close()
        return "Athlète introuvable", 404

    objectifs = q(conn, "SELECT * FROM objectifs WHERE athlete_id=%s ORDER BY epreuve", (aid,))
    resultats = q(conn, "SELECT * FROM resultats WHERE athlete_id=%s ORDER BY date_competition DESC", (aid,))
    champs = q(conn, "SELECT * FROM objectifs_champ WHERE athlete_id=%s ORDER BY epreuve", (aid,))
    conn.close()

    return render_template('admin_athlete.html', a=a, ath=a, objectifs=objectifs, resultats=resultats, champs=champs)

@app.route('/admin/minimas')
@login_required
@admin_required
def admin_minimas():
    """Show athletes who have reached CRF access thresholds this season."""
    conn = get_db()
    all_athletes = q(conn, """
        SELECT a.*, u.full_name as coach_name
        FROM athletes a JOIN users u ON a.coach_id=u.id
        ORDER BY a.categorie, a.nom_prenom
    """)
    conn.close()

    saison_courante = '2025/2026'
    resultats_saison = [r for r in RESULTATS_JSON if r.get('saison') == saison_courante]

    # Build best result per athlete per epreuve this season
    best_by_ath = {}  # licence -> {epreuve -> best_perf_seconds}
    for r in resultats_saison:
        lic = r['licence']
        ep = r.get('epreuve', '')
        perf_str = str(r.get('resultat') or r.get('performance') or '')
        perf_sec = parse_perf(perf_str, ep)
        if perf_sec is None: continue
        if lic not in best_by_ath: best_by_ath[lic] = {}
        if ep not in best_by_ath[lic] or perf_sec < best_by_ath[lic][ep]:
            best_by_ath[lic][ep] = (perf_sec, perf_str)

    results = []
    for a in all_athletes:
        lic = a['numero_licence']
        cat = a['categorie']
        sexe = a['sexe']
        bests = best_by_ath.get(lic, {})
        athlete_results = []
        for ep, minima_cats in MINIMAS.items():
            sexe_minimas = minima_cats.get(sexe, {})
            # Next cat minima
            next_cat = sexe_minimas.get(cat)
            if not next_cat: continue
            if ep not in bests: continue
            perf_sec, perf_str = bests[ep]
            min_sec = parse_perf(next_cat, ep)
            if min_sec is None: continue
            # For track: lower is better; for field: higher is better
            is_field = any(x in ep for x in ['Longueur','Triple','Hauteur','Poids','Disque','Javelot','Marteau'])
            if is_field:
                reached = perf_sec >= min_sec
            else:
                reached = perf_sec <= min_sec
            if reached:
                athlete_results.append({
                    'epreuve': ep, 'perf': perf_str,
                    'minima': next_cat, 'reached': True
                })
        if athlete_results:
            results.append({'athlete': a, 'results': athlete_results})

    return render_template('admin.html',
        minimas_results=results,
        minimas_saison=saison_courante,
        stats={'total_athletes': len(all_athletes), 'total_coaches': 0, 'en_attente': 0},
        pending_obj=[], pending_res=[], pending_champ=[], pending_perf=[],
        athletes=all_athletes, coaches=[], validated_obj=[], validated_res=[],
        active_admin_tab='tab-minimas'
    )

def parse_perf(perf_str, epreuve=''):
    """Convert performance string to seconds (or meters for field events)."""
    if not perf_str or perf_str in ('nan','None','—',''): return None
    s = str(perf_str).strip().upper().replace(',','.')
    import re as _re
    # Field events: remove M, KG, G suffix
    field_match = _re.match(r'^([0-9.]+)\s*(M|KG|G)$', s)
    if field_match:
        return float(field_match.group(1))
    # Time: convert to seconds
    s = s.lower()
    # apostrophe format: 8'23"56
    if "'" in s:
        s = s.replace("'",':').replace('"','.')
    # 3-part dot: 1.59.00
    dots = s.count('.')
    colons = s.count(':')
    if dots == 2 and colons == 0:
        parts = s.split('.')
        s = parts[0]+':'+parts[1]+'.'+parts[2]
    if ':' in s:
        parts = s.split(':')
        if len(parts) == 2:
            return float(parts[0])*60 + float(parts[1])
        if len(parts) == 3:
            h,m,sec = float(parts[0]),float(parts[1]),float(parts[2])
            if h<=2 and m<60 and sec<100:
                return h*60 + m + sec/100
            return h*3600 + m*60 + sec
    try: return float(s)
    except: return None

@app.route('/admin/export')
@login_required
@admin_required
def export_excel():
    # Get selected fields and filters
    fields = request.args.getlist('fields') or [
        'nom','licence','centre','entraineur','categorie','sexe','specialite','club','statut',
        'date_naissance','age','integration','classe',
        'obj_epreuve','obj_saison','obj_chrono','obj_statut','obj_date_validation',
        'res_epreuve','res_competition','res_lieu','res_date','res_classement','res_performance','res_date_validation',
        'hist_saison','hist_competition','hist_lieu','hist_date','hist_epreuve','hist_classement','hist_resultat'
    ]
    filter_coach = request.args.get('filter_coach','')
    filter_cat   = request.args.get('filter_cat','')
    filter_classe= request.args.get('filter_classe','')

    conn = get_db()

    # Athletes
    where = "WHERE 1=1"
    params = []
    if filter_coach:
        where += " AND u.full_name=%s"; params.append(filter_coach)
    if filter_cat:
        where += " AND a.categorie=%s"; params.append(filter_cat)
    if filter_classe:
        where += " AND a.classe=%s"; params.append(filter_classe)

    all_athletes = q(conn, f"""
        SELECT a.*, u.full_name as coach_name
        FROM athletes a JOIN users u ON a.coach_id=u.id
        {where} ORDER BY a.centre, a.categorie, a.nom_prenom
    """, params)

    # Objectifs validés
    all_obj = q(conn, """
        SELECT o.*, a.numero_licence
        FROM objectifs o JOIN athletes a ON o.athlete_id=a.id
        WHERE o.statut='validé' ORDER BY a.nom_prenom, o.epreuve
    """)
    obj_by_lic = {}
    for o in all_obj:
        obj_by_lic.setdefault(o['numero_licence'], []).append(o)

    # Résultats saisis validés
    all_res = q(conn, """
        SELECT r.*, a.numero_licence
        FROM resultats r JOIN athletes a ON r.athlete_id=a.id
        WHERE r.statut='validé' ORDER BY a.nom_prenom, r.date_competition
    """)
    res_by_lic = {}
    for r in all_res:
        res_by_lic.setdefault(r['numero_licence'], []).append(r)

    conn.close()

    # Historique JSON
    hist_by_lic = RESULTATS_BY_LICENCE

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Export AthléPro"

    # Build header map
    field_labels = {
        'nom':'Nom & Prénom','licence':'N° Licence','centre':'Centre/CRF',
        'entraineur':'Entraineur','categorie':'Catégorie','sexe':'Sexe',
        'specialite':'Spécialité','club':'Club','statut':'Statut',
        'date_naissance':'Date Naissance','age':'Âge','integration':'Date Intégration',
        'classe':'Classe',
        'obj_epreuve':'Obj. Épreuve','obj_saison':'Obj. Saison','obj_chrono':'Obj. Chrono',
        'obj_statut':'Obj. Statut','obj_date_validation':'Obj. Validé le',
        'res_epreuve':'Rés. Épreuve','res_competition':'Rés. Compétition','res_lieu':'Rés. Lieu',
        'res_date':'Rés. Date','res_classement':'Rés. Classement','res_performance':'Rés. Performance',
        'res_date_validation':'Rés. Validé le',
        'hist_saison':'Hist. Saison','hist_competition':'Hist. Compétition','hist_lieu':'Hist. Lieu',
        'hist_date':'Hist. Date','hist_epreuve':'Hist. Épreuve','hist_classement':'Hist. Classement',
        'hist_resultat':'Hist. Résultat',
    }

    # Styles
    hdr_font  = Font(bold=True, color='FFFFFF', size=9)
    hdr_fills = {
        'ath':  PatternFill(fill_type='solid', fgColor='1E2742'),
        'obj':  PatternFill(fill_type='solid', fgColor='1A6B3C'),
        'res':  PatternFill(fill_type='solid', fgColor='8B1A1A'),
        'hist': PatternFill(fill_type='solid', fgColor='B7500A'),
    }
    thin = Border(left=Side(style='thin',color='DDDDDD'), right=Side(style='thin',color='DDDDDD'),
                  bottom=Side(style='thin',color='DDDDDD'))
    alt  = PatternFill(fill_type='solid', fgColor='F8F9FC')
    ctr  = Alignment(horizontal='center', vertical='center')
    lft  = Alignment(vertical='center')

    def field_group(f):
        if f.startswith('obj_'):  return 'obj'
        if f.startswith('res_'):  return 'res'
        if f.startswith('hist_'): return 'hist'
        return 'ath'

    ath_fields  = [f for f in fields if field_group(f)=='ath']
    obj_fields  = [f for f in fields if field_group(f)=='obj']
    res_fields  = [f for f in fields if field_group(f)=='res']
    hist_fields = [f for f in fields if field_group(f)=='hist']
    need_obj    = bool(obj_fields)
    need_res    = bool(res_fields)
    need_hist   = bool(hist_fields)

    # Write header
    col = 1
    for f in fields:
        cell = ws.cell(row=1, column=col, value=field_labels.get(f, f))
        cell.font = hdr_font
        cell.fill = hdr_fills[field_group(f)]
        cell.alignment = ctr
        cell.border = thin
        col += 1
    ws.row_dimensions[1].height = 22

    # Write data rows
    data_row = 2
    for a in all_athletes:
        lic    = a['numero_licence']
        objs   = obj_by_lic.get(lic, [None])
        ress   = res_by_lic.get(lic, [None])
        hists  = hist_by_lic.get(lic, [None])
        n_rows = max(
            len(objs)  if need_obj  else 1,
            len(ress)  if need_res  else 1,
            len(hists) if need_hist else 1,
            1
        )
        bg = alt if data_row % 2 == 0 else None

        for i in range(n_rows):
            obj  = objs[i]  if (need_obj  and i < len(objs))  else {}
            res  = ress[i]  if (need_res  and i < len(ress))  else {}
            hist = hists[i] if (need_hist and i < len(hists)) else {}

            field_vals = {
                'nom':  a['nom_prenom']  if i==0 else '',
                'licence':  a['numero_licence']  if i==0 else '',
                'centre':   a['centre']           if i==0 else '',
                'entraineur': a['coach_name']     if i==0 else '',
                'categorie':  a['categorie']      if i==0 else '',
                'sexe':       a['sexe']            if i==0 else '',
                'specialite': a['specialite']      if i==0 else '',
                'club':       a['club']            if i==0 else '',
                'statut':     a['statut']          if i==0 else '',
                'date_naissance': (str(a['date_naissance'])[:10] if a['date_naissance'] else '') if i==0 else '',
                'age':        str(a['age'] or '')  if i==0 else '',
                'integration': (str(a['date_integration'])[:10] if a['date_integration'] else '') if i==0 else '',
                'classe':     a.get('classe','')   if i==0 else '',
                'obj_epreuve':         (obj or {}).get('epreuve',''),
                'obj_saison':          (obj or {}).get('saison',''),
                'obj_chrono':          (obj or {}).get('objectif_chrono',''),
                'obj_statut':          (obj or {}).get('statut',''),
                'obj_date_validation': str((obj or {}).get('valide_le','') or '')[:10],
                'res_epreuve':         (res or {}).get('epreuve',''),
                'res_competition':     (res or {}).get('nom_competition',''),
                'res_lieu':            (res or {}).get('lieu',''),
                'res_date':            str((res or {}).get('date_competition','') or '')[:10],
                'res_classement':      str((res or {}).get('classement','') or ''),
                'res_performance':     (res or {}).get('performance',''),
                'res_date_validation': str((res or {}).get('valide_le','') or '')[:10],
                'hist_saison':      (hist or {}).get('saison',''),
                'hist_competition': (hist or {}).get('competition',''),
                'hist_lieu':        (hist or {}).get('lieu',''),
                'hist_date':        str((hist or {}).get('date','') or '')[:10],
                'hist_epreuve':     (hist or {}).get('epreuve',''),
                'hist_classement':  str((hist or {}).get('classement','') or ''),
                'hist_resultat':    str((hist or {}).get('resultat','') or ''),
            }

            for col_idx, f in enumerate(fields, 1):
                cell = ws.cell(row=data_row, column=col_idx, value=field_vals.get(f,''))
                cell.border = thin
                cell.alignment = lft
                if bg: cell.fill = bg

            data_row += 1

    # Column widths
    for col_idx, f in enumerate(fields, 1):
        w = 22 if f in ('nom','centre','entraineur','res_competition','hist_competition') else 14
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.freeze_panes = 'A2'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name='athletepro_export.xlsx',
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── AUTO INIT ──
init_db()
