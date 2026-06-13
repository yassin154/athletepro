#!/usr/bin/env python3
"""
export_rankings.py — AthléPro
Exporte wa_rankings.json en Excel pour vérification.
Usage: python export_rankings.py
"""

import json, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

with open(os.path.join(BASE_DIR, 'wa_rankings.json'), encoding='utf-8') as f:
    rankings = json.load(f)
with open(os.path.join(BASE_DIR, 'athletes.json'), encoding='utf-8') as f:
    athletes = json.load(f)

# Build licence -> nom map
lic_to_nom = {a['licence']: a['nom'].strip() for a in athletes}

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "World Rankings MAR"

# Header
headers = ['Licence', 'Nom', 'Épreuve', 'Rang Mondial', 'Score', 'Mark', 'Date']
hdr_font = Font(bold=True, color='FFFFFF', size=10)
hdr_fill = PatternFill(fill_type='solid', fgColor='1E2742')
ctr = Alignment(horizontal='center', vertical='center')

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = ctr
ws.row_dimensions[1].height = 20

# Data
alt_fill = PatternFill(fill_type='solid', fgColor='F0F5FF')
row_idx = 2
for lic, ranks in sorted(rankings.items()):
    nom = lic_to_nom.get(lic, f'Licence {lic}')
    for r in sorted(ranks, key=lambda x: (x.get('discipline',''), str(x.get('rank_int') or 9999))):
        bg = alt_fill if row_idx % 2 == 0 else None
        vals = [
            lic,
            nom,
            r.get('discipline',''),
            r.get('rank_int'),
            r.get('score',''),
            r.get('mark',''),
            r.get('date',''),
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment = Alignment(vertical='center')
            if bg: cell.fill = bg
            if col == 4 and val:  # Rang mondial en bleu
                cell.font = Font(color='2471D6', bold=True)
        row_idx += 1

# Widths
for col, w in enumerate([14, 28, 20, 14, 10, 12, 12], 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
ws.freeze_panes = 'A2'

out = os.path.join(BASE_DIR, 'wa_rankings_export.xlsx')
wb.save(out)
print(f"✅ Export: {out}")
print(f"   {row_idx-2} lignes | {len(rankings)} athlètes")
